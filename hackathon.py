from flask import Flask, request, jsonify
from flask_cors import CORS
from bs4 import BeautifulSoup
import requests
import PyPDF2
import docx
import openpyxl
from groq import Groq
import pyaudio
import wave
import pyttsx3
import threading
import os
#from together import Together

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB
CORS(app)

API_KEY = os.getenv('GROQ_API_KEY', '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')  # Use an environment variable
client = Groq(api_key=API_KEY)

class TextToSpeechEngine:
    def __init__(self):
        self.engine = pyttsx3.init()
        self.engine.setProperty('voice', 'HKEY_LOCAL_MACHINE\\SOFTWARE\\Microsoft\\Speech\\Voices\\Tokens\\TTS_MS_EN-US_ZIRA_11.0')
        self.engine.setProperty('rate', 150)

    def read_text(self, text):
        if self.engine._inLoop:
            self.engine.endLoop()
        self.engine.say(text)
        self.engine.runAndWait()
        if self.engine._inLoop:
            self.engine.endLoop()

class AudioRecorder:
    def __init__(self):
        self.audio = pyaudio.PyAudio()
        self.stream = None
        self.frames = []

    def start_recording(self):
        self.stream = self.audio.open(format=pyaudio.paInt16, channels=2, rate=44100, input=True, frames_per_buffer=1024)
        self.frames = []
        print("Recording...")

    def end_recording(self, filename):
        self.stream.stop_stream()
        self.stream.close()
        self.audio.terminate()
        with wave.open(filename, 'wb') as waveFile:
            waveFile.setnchannels(2)
            waveFile.setsampwidth(self.audio.get_sample_size(pyaudio.paInt16))
            waveFile.setframerate(44100)
            waveFile.writeframes(b''.join(self.frames))
        print(f"Finished recording and saved to {filename}")

class FileProcessor:
    def process_file(self, file):
        processors = {
            '.pdf': self.process_pdf,
            '.docx': self.process_docx,
            '.xlsx': self.process_xlsx
        }
        ext = os.path.splitext(file.filename)[1].lower()
        if ext in processors:
            return processors[ext](file)
        else:
            raise ValueError('Unsupported file type')

    def process_pdf(self, file):
        pdf_file = PyPDF2.PdfReader(file)
        return ''.join(page.extract_text() or '' for page in pdf_file.pages)

    def process_docx(self, file):
        doc = docx.Document(file)
        return ''.join(para.text for para in doc.paragraphs)

    def process_xlsx(self, file):
        workbook = openpyxl.load_workbook(file)
        return ' '.join(' '.join(map(str, row)) for sheet in workbook.sheetnames for row in workbook[sheet].iter_rows(values_only=True))

class SummaryGenerator:
    def generate_summary(self, prompt, summary_options):
        prompt2 = f"summarize the below. The prompt is: {prompt}"
        if summary_options:
            prompt2 += " " + " ".join(summary_options)
        
        chat_completion = client.chat.completions.create(
            messages=[{"role": "user", "content": prompt2}],
            model="llama-3.1-70b-versatile",
        )
        return chat_completion.choices[0].message.content

text_to_speech_engine = TextToSpeechEngine()
audio_recorder = AudioRecorder()
file_processor = FileProcessor()
summary_generator = SummaryGenerator()

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json
    prompt = data.get('prompt')
    if not prompt:
        return jsonify({'error': 'Prompt is required'}), 400

    summary_options = [
        f"in {data.get('summaryType')} format",
        f"in {data.get('summaryLength')} length",
        f"using {data.get('formatting')}"
    ]
    if data.get('conclusion'):
        summary_options.append("Please include a conclusion.")
    if data.get('academic'):
        summary_options.append("Use an academic tone.")
    if data.get('title'):
        summary_options.append("Provide a title for the summary.")

    response = summary_generator.generate_summary(prompt, summary_options)
    return jsonify({'response': response})

@app.route('/summarize_url', methods=['POST'])
def summarize_url():
    url = request.json.get('url')
    if not url:
        return jsonify({'error': 'URL is required'}), 400

    try:
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')
        text = soup.get_text()
        return jsonify({'response': summary_generator.generate_summary(text, [])})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file provided'}), 400

    try:
        text = file_processor.process_file(file)
        response = summary_generator.generate_summary(text, [])
        return jsonify({'response': response})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/read', methods=['POST'])
def read_text():
    text = request.json['text']
    threading.Thread(target=text_to_speech_engine.read_text, args=(text,)).start()
    return jsonify({'message': 'Text read successfully'})

if __name__ == '__main__':
    app.run(debug=True)